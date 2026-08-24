from __future__ import annotations

import math
import os
import re
import shutil
import tempfile
import warnings
from typing import Any

import fitz
import openpyxl


# ============================================================
# CONFIGURATION DEFAULTS
# ============================================================

RED = (1.0, 0.0, 0.0)

CIRCLE_SCALE = 1.9
CIRCLE_MAX_DIAMETER = 11.0       # Test Point (non-PAX) rings
CIRCLE_MAX_DIAMETER_PAX = 11.0   # Connector (PAX) rings
CIRCLE_LINE_WIDTH = 0.9
# Per-category ring offsets: PAX = Connector circles, TP = Test Point (non-PAX) circles.
# Nudges the ring (and its label anchor) off the detected test point.
CIRCLE_OFFSET_X_PAX = 0.0
CIRCLE_OFFSET_Y_PAX = 0.0
CIRCLE_OFFSET_X_TP = 0.0
CIRCLE_OFFSET_Y_TP = 0.0

LABEL_FONT_SIZE = 4.5
LABEL_FONT_NAME = "helv"
LABEL_WRAP_CHARS = 7
LABEL_LINE_SPACING = 0.7

LABEL_RADIAL_OFFSET = 5.0  # Base distance pushed away from circle (Test Point AND Connector labels)

# Test Point (non-connector) labels: independent offset + attachment side.
# LABEL_ATTACH_TP: "top" | "bottom" | "left" | "right" (which side of the ring the label attaches to).
LABEL_OFFSET_X_TP = 0.0
LABEL_OFFSET_Y_TP = 0.0
LABEL_ATTACH_TP = "top"

# Excel routing matrix: nets in spreadsheet rows >= this are forced to the slave PDF only
SLAVE_ONLY_FROM_ROW = 460

# Advanced PAX specific offsets (Dynamic Split)
LABEL_OFFSET_X_PAX_FIRST_HALF = 0.0
LABEL_OFFSET_Y_PAX_FIRST_HALF = 0.0

LABEL_OFFSET_X_PAX_SECOND_HALF = 0.0
LABEL_OFFSET_Y_PAX_SECOND_HALF = 0.0

LABEL_BG_ENABLED = True
LABEL_BG_COLOR = (1.0, 1.0, 1.0)
LABEL_BG_COLOR_HEX = "#ffffff"
LABEL_BG_ALPHA = 1.0
LABEL_BG_PADDING_X = 0.2
LABEL_BG_PADDING_Y = 0.2


# ============================================================
# BASIC HELPERS
# ============================================================

def _clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().upper().replace(" ", "")

def _is_invalid(value: str) -> bool:
    return _clean(value) in {
        "", "N/A", "NA", "-", "—",
        "TESTPOINT", "TESTPOINTS", "TESTPOINT.", "TESTPOINT:",
    }

def parse_hex_color(hex_color: str) -> tuple[float, float, float]:
    """Convert a #RGB or #RRGGBB string to normalized RGB tuple."""
    s = str(hex_color).strip().lstrip("#")
    if len(s) == 3:
        s = "".join(ch * 2 for ch in s)
    if len(s) != 6:
        raise ValueError("Hex color must be #RRGGBB or #RGB")
    return (
        int(s[0:2], 16) / 255.0,
        int(s[2:4], 16) / 255.0,
        int(s[4:6], 16) / 255.0,
    )

def get_label_bg_color() -> tuple[float, float, float]:
    if LABEL_BG_COLOR_HEX:
        try:
            return parse_hex_color(LABEL_BG_COLOR_HEX)
        except Exception:
            pass
    return LABEL_BG_COLOR


def _coerce_float(value: Any, fallback: float) -> float:
    try:
        parsed = float(value)
        if parsed != parsed or parsed in (float("inf"), float("-inf")):
            return fallback
        return parsed
    except (TypeError, ValueError):
        return fallback


def _coerce_int(value: Any, fallback: int) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return fallback


def _coerce_attach(value: Any, fallback: str) -> str:
    normalized = str(value).strip().lower()
    return normalized if normalized in {"top", "bottom", "left", "right"} else fallback


def _coerce_hex(value: Any, fallback: str) -> str:
    raw = str(value).strip()
    if not raw:
        return fallback
    cleaned = raw if raw.startswith("#") else f"#{raw}"
    try:
        parse_hex_color(cleaned)
        return cleaned
    except ValueError:
        return fallback


# ============================================================
# DYNAMIC CONFIGURATION LOADER (CALLED BY FLASK)
# ============================================================

def apply_config(cfg: dict | None) -> None:
    """Apply runtime UI configuration safely, preserving current values on invalid input."""
    global RED, CIRCLE_SCALE, CIRCLE_MAX_DIAMETER, CIRCLE_MAX_DIAMETER_PAX, CIRCLE_LINE_WIDTH
    global CIRCLE_OFFSET_X_PAX, CIRCLE_OFFSET_Y_PAX, CIRCLE_OFFSET_X_TP, CIRCLE_OFFSET_Y_TP
    global LABEL_FONT_SIZE, LABEL_WRAP_CHARS, LABEL_LINE_SPACING
    global LABEL_RADIAL_OFFSET
    global LABEL_OFFSET_X_PAX_FIRST_HALF, LABEL_OFFSET_Y_PAX_FIRST_HALF
    global LABEL_OFFSET_X_PAX_SECOND_HALF, LABEL_OFFSET_Y_PAX_SECOND_HALF
    global LABEL_OFFSET_X_TP, LABEL_OFFSET_Y_TP
    global LABEL_ATTACH_TP
    global LABEL_BG_ENABLED, LABEL_BG_COLOR_HEX, LABEL_BG_COLOR, LABEL_BG_ALPHA
    global LABEL_BG_PADDING_X, LABEL_BG_PADDING_Y

    if not cfg:
        return

    if "RED" in cfg:
        val = cfg["RED"]
        if isinstance(val, str):
            try:
                RED = parse_hex_color(val)
            except ValueError:
                pass
        elif isinstance(val, (list, tuple)) and len(val) == 3:
            try:
                RED = tuple(float(x) for x in val)
            except (TypeError, ValueError):
                pass

    if "LABEL_BG_COLOR_HEX" in cfg:
        LABEL_BG_COLOR_HEX = _coerce_hex(cfg["LABEL_BG_COLOR_HEX"], LABEL_BG_COLOR_HEX)
    if "LABEL_BG_ALPHA" in cfg:
        LABEL_BG_ALPHA = max(0.0, min(1.0, _coerce_float(cfg["LABEL_BG_ALPHA"], LABEL_BG_ALPHA)))
    if "LABEL_BG_ENABLED" in cfg:
        LABEL_BG_ENABLED = bool(cfg["LABEL_BG_ENABLED"])
    
    for key, target in [
        ("CIRCLE_SCALE", "CIRCLE_SCALE"),
        ("CIRCLE_MAX_DIAMETER", "CIRCLE_MAX_DIAMETER"),
        ("CIRCLE_MAX_DIAMETER_PAX", "CIRCLE_MAX_DIAMETER_PAX"),
        ("CIRCLE_LINE_WIDTH", "CIRCLE_LINE_WIDTH"),
        ("CIRCLE_OFFSET_X_PAX", "CIRCLE_OFFSET_X_PAX"),
        ("CIRCLE_OFFSET_Y_PAX", "CIRCLE_OFFSET_Y_PAX"),
        ("CIRCLE_OFFSET_X_TP", "CIRCLE_OFFSET_X_TP"),
        ("CIRCLE_OFFSET_Y_TP", "CIRCLE_OFFSET_Y_TP"),
        ("LABEL_FONT_SIZE", "LABEL_FONT_SIZE"),
        ("LABEL_WRAP_CHARS", "LABEL_WRAP_CHARS"),
        ("LABEL_LINE_SPACING", "LABEL_LINE_SPACING"),
        ("LABEL_RADIAL_OFFSET", "LABEL_RADIAL_OFFSET"),
        ("LABEL_OFFSET_X_PAX_FIRST_HALF", "LABEL_OFFSET_X_PAX_FIRST_HALF"),
        ("LABEL_OFFSET_Y_PAX_FIRST_HALF", "LABEL_OFFSET_Y_PAX_FIRST_HALF"),
        ("LABEL_OFFSET_X_PAX_SECOND_HALF", "LABEL_OFFSET_X_PAX_SECOND_HALF"),
        ("LABEL_OFFSET_Y_PAX_SECOND_HALF", "LABEL_OFFSET_Y_PAX_SECOND_HALF"),
        ("LABEL_OFFSET_X_TP", "LABEL_OFFSET_X_TP"),
        ("LABEL_OFFSET_Y_TP", "LABEL_OFFSET_Y_TP"),
    ]:
        if key in cfg:
            current = globals()[target]
            if key == "LABEL_WRAP_CHARS":
                globals()[target] = max(0, _coerce_int(cfg[key], int(current)))
            else:
                globals()[target] = _coerce_float(cfg[key], float(current))

    for key in ("LABEL_ATTACH_TP",):
        if key in cfg:
            globals()[key] = _coerce_attach(cfg[key], globals()[key])


# ============================================================
# EXCEL PROJECT DETECTION AND PARSING
# ============================================================

def get_sheet_names(xlsx_path: str) -> list[str]:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    return wb.sheetnames

def get_sheet_headers(xlsx_path: str, sheet_name: str) -> list[str]:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]
    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(1, col).value
        headers.append(str(val).strip() if val else f"Col {col}")
    return headers

def detect_projects(xlsx_path: str, sheet_name: str = "overview DSUB") -> list[str]:
    """Detect project identifiers from the overview sheet header region."""
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    projects: list[str] = []
    banned = {"NET NAME", "TEST POINT", "TESTPOINT", "PIN DSUB", "PIN HARTING", "CHANNEL VT", "CHANNEL DUT", "CABLE TYPE", "WIRE"}

    for row in range(1, min(ws.max_row, 25) + 1):
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row, col).value
            if not val:
                continue
            text = str(val).strip()
            upper = text.upper()
            if upper in banned:
                continue
            if ("BMW" in upper or "JLR" in upper or "MERCEDES" in upper or upper.startswith("PLATFORM") or re.search(r"V\d+A\.\d+\.\d+", upper)):
                if text not in projects:
                    projects.append(text)
    return projects

def _find_project_column_pairs(ws, project_name: str) -> list[tuple[int, int]]:
    target = str(project_name).strip().upper()
    pairs: list[tuple[int, int]] = []
    for row in range(1, min(ws.max_row, 60) + 1):
        for col in range(1, ws.max_column):
            val = ws.cell(row, col).value
            if not val or str(val).strip().upper() != target:
                continue
            for header_row in range(row + 1, min(ws.max_row, row + 20) + 1):
                left = ws.cell(header_row, col).value
                right = ws.cell(header_row, col + 1).value
                left_s = str(left).strip().upper() if left else ""
                right_s = str(right).strip().upper() if right else ""
                if left_s == "NET NAME" and right_s in {"TEST POINT", "TESTPOINT"}:
                    candidate = (col, col + 1)
                    if candidate not in pairs:
                        pairs.append(candidate)
                    break
    return pairs

def _valid_label(label: str) -> bool:
    s = _clean(label)
    if _is_invalid(s):
        return False
    patterns = [r"T-?\d+", r"TP-?\d+", r"X\d+[.-]\d+", r"PAX\d+", r"PATP\d+", r"COTP\d+", r"GND"]
    return any(re.fullmatch(p, s) for p in patterns)

def excel_tp_to_pdf_labels(tp: str) -> list[str]:
    if tp is None:
        return []
    raw = str(tp).strip().upper().replace(" ", "")
    if _is_invalid(raw):
        return []
    results: list[str] = []
    grouped = re.fullmatch(r"(X\d+)([-.])(\d+(?:,\d+)*)", raw)
    if grouped:
        prefix = grouped.group(1)
        sep = grouped.group(2)
        for n in grouped.group(3).split(","):
            candidate = f"{prefix}{sep}{int(n)}"
            if _valid_label(candidate):
                results.append(candidate)
        return list(dict.fromkeys(results))

    for part in re.split(r"[;/]+", raw):
        part = part.strip()
        if not part or _is_invalid(part):
            continue
        if _valid_label(part):
            results.append(part)
    return list(dict.fromkeys(results))

def read_excel_data(xlsx_path: str, include_s_nets: bool = True, net_col: int = 25, tp_col: int = 26, sheet_name: str = "overview DSUB", project_name: str | None = None) -> tuple[list[tuple[str, str]], list[dict]]:
    """Read and normalize (test-point, net) pairs from the selected project columns."""
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found")
    ws = wb[sheet_name]
    column_pairs = _find_project_column_pairs(ws, project_name) if project_name else [(net_col, tp_col)]
    if project_name and not column_pairs:
        raise ValueError(f"Could not find Net Name/Test point columns for project '{project_name}'")

    # Each result is (label, net, row): row is the source spreadsheet row, used so that
    # nets from row >= SLAVE_ONLY_FROM_ROW can be forced to the slave PDF in categorize_pairs.
    results: list[tuple[str, str, int]] = []
    skipped: list[dict] = []
    seen: set[tuple[str, str]] = set()

    for net_c, tp_c in column_pairs:
        for row in range(1, ws.max_row + 1):
            net_raw = ws.cell(row, net_c).value
            tp_raw = ws.cell(row, tp_c).value
            net = str(net_raw).strip() if net_raw is not None else ""
            tp = str(tp_raw).strip() if tp_raw is not None else ""

            if _clean(net) in {"", "NETNAME"} and _clean(tp) in {"", "TESTPOINT", "TESTPOINTS"}:
                continue
            if _clean(tp) in {"", "TESTPOINT", "TESTPOINTS", "N/A", "NA", "-"}:
                continue

            labels = excel_tp_to_pdf_labels(tp)
            if not labels:
                if net or tp:
                    skipped.append({"row": row, "tp_raw": tp, "net_raw": net, "reason": "No valid test point"})
                continue
            for label in labels:
                key = (label, net)
                if key in seen:
                    continue
                seen.add(key)
                results.append((label, net, row))
    return results, skipped

def categorize_pairs(pairs: list[tuple], skipped: list[dict]) -> dict:
    master: list[dict] = []
    slave: list[dict] = []
    for entry in pairs:
        # Accept both (label, net) and (label, net, row) for backward compatibility
        pdf_label, net_label = entry[0], entry[1]
        row = entry[2] if len(entry) > 2 else None
        net_u = str(net_label).strip().upper()
        item = {"pdf_label": pdf_label, "net_label": net_label}
        # Nets from row >= SLAVE_ONLY_FROM_ROW are slave-only regardless of net prefix
        if (row is not None and row >= SLAVE_ONLY_FROM_ROW) or net_u.startswith("S-") or net_u.startswith("SLAVE"):
            slave.append(item)
        else:
            master.append(item)
    return {"master": master, "slave": slave, "skipped": skipped}


# ============================================================
# PDF-FIRST DYNAMIC SCANNING & ALIASES
# ============================================================

def detect_dominant_prefix(page_words: list, page_text: str) -> tuple[str, dict]:
    counts = {"T": 0, "TP": 0, "TP-": 0, "COT": 0, "COTP": 0, "PATP": 0}
    for w in page_words:
        w_text = str(w[4]).strip().upper()
        if re.fullmatch(r"T\d+", w_text): counts["T"] += 1
        elif re.fullmatch(r"TP\d+", w_text): counts["TP"] += 1
        elif re.fullmatch(r"TP-\d+", w_text): counts["TP-"] += 1
        elif re.fullmatch(r"COT\d+", w_text): counts["COT"] += 1
        elif re.fullmatch(r"COTP\d+", w_text): counts["COTP"] += 1
        elif re.fullmatch(r"PATP\d+", w_text): counts["PATP"] += 1
        
    counts["TP "] = len(re.findall(r"\bTP \d+\b", page_text.upper()))
    
    c_counts = {k: v for k, v in counts.items() if k.startswith("C")}
    best_c = max(c_counts, key=c_counts.get)
    
    if c_counts[best_c] > 0:
        return best_c, counts
        
    best = max(counts, key=counts.get)
    return (best if counts[best] > 0 else "T"), counts


def build_aliases(label: str, dominant_prefix: str = "T") -> list[str]:
    raw = _clean(label)
    if not raw:
        return []
        
    m = re.fullmatch(r"([A-Z]+)[-_]?(\d+)", raw)
    if m:
        prefix = m.group(1)
        if prefix not in ("PAX", "X"):
            n = str(int(m.group(2)))  # Strip leading zeros
            if dominant_prefix == "TP ":
                return [f"TP {n}"]
            return [f"{dominant_prefix}{n}"]

    aliases: list[str] = [raw]
    m = re.fullmatch(r"X(\d+)[.\-_/:]?(\d+)", raw)
    if m:
        a = int(m.group(1))
        b = int(m.group(2))
        aliases.extend([
            f"X{a}-{b}", f"X{a}.{b}", f"X{a}{b}", f"X{a}_{b}", f"X{a}/{b}", f"X{a}:{b}",
            f"PAX{a}{b:02d}", f"PAX{a}{b:03d}", f"PAX{a}-{b}", f"PAX{a}.{b}"
        ])
        if a == 3:
            aliases.extend([f"PAX2{b:02d}", f"PAX2{b:03d}", f"PAX2-{b}", f"PAX2.{b}"])

    m = re.fullmatch(r"PAX(\d+)", raw)
    if m:
        digits = m.group(1)
        if digits.startswith("3") and len(digits) >= 3:
            aliases.append("PAX2" + digits[1:])

    return list(dict.fromkeys(aliases))

def _word_rect(word) -> fitz.Rect:
    return fitz.Rect(word[0], word[1], word[2], word[3])

def _merge_words(words: list) -> list[tuple[str, fitz.Rect]]:
    merged: list[tuple[str, fitz.Rect]] = []
    words_sorted = sorted(words, key=lambda w: (round(w[1], 1), w[0]))
    for i in range(len(words_sorted)):
        text = ""
        rect = None
        base_y = words_sorted[i][1]
        for j in range(i, min(i + 15, len(words_sorted))):
            if abs(words_sorted[j][1] - base_y) > 15.0:
                break
            text += _clean(words_sorted[j][4])
            r = _word_rect(words_sorted[j])
            rect = r if rect is None else rect | r
            merged.append((text, rect))
    return merged

def _find_pin_by_proximity(page: fitz.Page, page_words: list, prefix: str, pin: str, merged_words: list | None = None) -> fitz.Rect | None:
    target_prefixes = [prefix]
    if prefix == "X3":
        target_prefixes.extend(["PAX2", "X2"])

    prefix_rects = []
    for w in page_words:
        if _clean(w[4]) in target_prefixes:
            prefix_rects.append(_word_rect(w))

    if not prefix_rects:
        merged = merged_words if merged_words is not None else _merge_words(page_words)
        for text, r in merged:
            if text in target_prefixes:
                prefix_rects.append(r)
                
    if not prefix_rects:
        return None

    pin_rects = []
    for w in page_words:
        if _clean(w[4]) == pin:
            pin_rects.append(_word_rect(w))
    if not pin_rects:
        return None

    best_rect = None
    min_dist = float('inf')
    MAX_DIST = 1500.0  

    for pr in prefix_rects:
        cx1 = (pr.x0 + pr.x1) / 2.0
        cy1 = (pr.y0 + pr.y1) / 2.0
        for pir in pin_rects:
            cx2 = (pir.x0 + pir.x1) / 2.0
            cy2 = (pir.y0 + pir.y1) / 2.0
            dist = ((cx1 - cx2)**2 + (cy1 - cy2)**2)**0.5
            if dist < min_dist and dist < MAX_DIST:
                min_dist = dist
                best_rect = pir
    return best_rect

def find_tp_rect(page: fitz.Page, pdf_label: str, page_words: list, dominant_prefix: str,
                 aliases: list | None = None, cleaned_words: list | None = None,
                 merged_words: list | None = None) -> fitz.Rect | None:
    if aliases is None:
        aliases = build_aliases(pdf_label, dominant_prefix)
    if cleaned_words is None:
        cleaned_words = [(_clean(w[4]), w) for w in page_words]

    for alias in aliases:
        target = _clean(alias)
        for ctext, w in cleaned_words:
            if ctext == target:
                return _word_rect(w)

    if merged_words is None:
        merged_words = _merge_words(page_words)
    for alias in aliases:
        target = _clean(alias)
        for text, rect in merged_words:
            if text == target:
                return rect

    for alias in aliases:
        rects = page.search_for(alias)
        for r in rects:
            context = _clean(page.get_textbox(r + (-2, -2, 2, 2)))
            if context == _clean(alias):
                return r

    m = re.fullmatch(r"(X\d+)[.\-_/:]?(\d+)", _clean(pdf_label))
    if m:
        prox_rect = _find_pin_by_proximity(page, page_words, m.group(1), m.group(2), merged_words=merged_words)
        if prox_rect:
            return prox_rect

    return None


# ============================================================
# SPIRAL SMART PLACEMENT & VISUALS
# ============================================================

def draw_leader_line(page: fitz.Page, label_cx: float, label_cy: float, target_cx: float, target_cy: float, circle_radius: float, color: tuple):
    dx = target_cx - label_cx
    dy = target_cy - label_cy
    dist = math.hypot(dx, dy)

    if dist <= circle_radius + 1.0:
        return

    ratio = (dist - circle_radius) / dist
    edge_x = label_cx + dx * ratio
    edge_y = label_cy + dy * ratio

    page.draw_line(
        fitz.Point(label_cx, label_cy),
        fitz.Point(edge_x, edge_y),
        color=color,
        width=0.4,
        dashes="[1 1]"
    )

def find_whitespace_spiral(cx: float, cy: float, label_w: float, label_h: float, preferred_angle_deg: float, occupied_labels: list[fitz.Rect], base_radius: float, padding: float=1.5):
    step_angle = 15  
    step_radius = 2.0 
    max_attempts = 300 

    for i in range(max_attempts):
        current_radius = base_radius + (i // (360 // step_angle)) * step_radius
        current_angle_deg = preferred_angle_deg + (i * step_angle)
        rad = math.radians(current_angle_deg % 360)

        cand_cx = cx + current_radius * math.cos(rad)
        cand_cy = cy + current_radius * math.sin(rad)

        cand_rect = fitz.Rect(
            cand_cx - (label_w / 2.0) - padding,
            cand_cy - (label_h / 2.0) - padding,
            cand_cx + (label_w / 2.0) + padding,
            cand_cy + (label_h / 2.0) + padding
        )

        collision = False
        for occ in occupied_labels:
            if not (cand_rect & occ).is_empty:
                collision = True
                break
        
        if not collision:
            return cand_cx, cand_cy, cand_rect

    fallback_rad = math.radians(preferred_angle_deg)
    fallback_radius = base_radius + (max_attempts // (360 // step_angle)) * step_radius
    fallback_cx = cx + fallback_radius * math.cos(fallback_rad)
    fallback_cy = cy + fallback_radius * math.sin(fallback_rad)
    
    fallback_rect = fitz.Rect(
        fallback_cx - (label_w / 2.0) - padding,
        fallback_cy - (label_h / 2.0) - padding,
        fallback_cx + (label_w / 2.0) + padding,
        fallback_cy + (label_h / 2.0) + padding
    )
    return fallback_cx, fallback_cy, fallback_rect

def circle_test_point(page: fitz.Page, tp_rect: fitz.Rect, pdf_label: str) -> tuple[float, float, float, fitz.Rect]:
    # Connector (PAX) vs Test Point (non-PAX) get independent ring offsets and max diameters
    is_pax = _normalize_label_for_visuals(pdf_label).startswith("PAX")
    off_x = CIRCLE_OFFSET_X_PAX if is_pax else CIRCLE_OFFSET_X_TP
    off_y = CIRCLE_OFFSET_Y_PAX if is_pax else CIRCLE_OFFSET_Y_TP
    max_diameter = CIRCLE_MAX_DIAMETER_PAX if is_pax else CIRCLE_MAX_DIAMETER
    cx = (tp_rect.x0 + tp_rect.x1) / 2.0 + off_x
    cy = (tp_rect.y0 + tp_rect.y1) / 2.0 + off_y

    base = max(tp_rect.width, tp_rect.height)
    scaled_radius = (base * CIRCLE_SCALE) / 2.0
    radius = max(0.8, min(max_diameter / 2.0, scaled_radius))

    shape = page.new_shape()
    shape.draw_circle((cx, cy), radius)
    shape.finish(color=RED, width=CIRCLE_LINE_WIDTH)
    shape.commit()

    circle_rect = fitz.Rect(cx - radius, cy - radius, cx + radius, cy + radius)
    return cx, cy, radius, circle_rect

def wrap_label_text(text: str, max_chars: int | None = None) -> str:
    # max_chars=None means "use current runtime config" rather than import-time default.
    if max_chars is None:
        max_chars = LABEL_WRAP_CHARS
    try:
        max_chars = int(max_chars)
    except (TypeError, ValueError):
        max_chars = LABEL_WRAP_CHARS

    s = str(text).strip()
    if not s or max_chars <= 0:
        return s
    lines: list[str] = []
    remaining = s
    while len(remaining) > max_chars:
        chunk = remaining[:max_chars + 1]
        split_at = chunk.rfind(" ")
        if split_at <= 0:
            split_at = max_chars
        line = remaining[:split_at].rstrip()
        if line:
            lines.append(line)
        remaining = remaining[split_at:].lstrip()
    if remaining:
        lines.append(remaining)
    return "\n".join(lines)

def make_label_rect(x0: float, x1: float, anchor_y: float, text_h: float, align_v: str) -> fitz.Rect:
    mode = (align_v or "bottom").strip().lower()
    if mode == "top":
        y0 = anchor_y
        y1 = anchor_y + text_h
    elif mode == "middle":
        y0 = anchor_y - (text_h / 2.0)
        y1 = anchor_y + (text_h / 2.0)
    else:
        y0 = anchor_y - text_h
        y1 = anchor_y
    return fitz.Rect(x0, y0, x1, y1)

def _normalize_label_for_visuals(label: str) -> str:
    s = _clean(label)
    m = re.fullmatch(r"X(\d+)[.\-_/:]?(\d+)", s)
    if m:
        return f"PAX{m.group(1)}{int(m.group(2)):02d}"
    return s

def draw_net_label_top_touch(
    page: fitz.Page, text: str, pdf_label: str, cx: float, cy: float, 
    radius: float, occupied_labels: list[fitz.Rect], pax_max_pins: dict | None = None,
    manual_dx: float = 0.0, manual_dy: float = 0.0
) -> fitz.Rect:
    wrapped = wrap_label_text(text)
    lines = wrapped.split("\n")
    font = fitz.Font(LABEL_FONT_NAME)
    line_h = LABEL_FONT_SIZE * (font.ascender - font.descender)
    line_step = line_h * LABEL_LINE_SPACING
    text_h = line_step * len(lines)
    text_w = max(fitz.get_text_length(line, fontname=LABEL_FONT_NAME, fontsize=LABEL_FONT_SIZE) for line in lines)
    
    check_label = _normalize_label_for_visuals(pdf_label)
    is_pax = check_label.startswith("PAX")

    if is_pax:
        is_first_half = True
        m = re.fullmatch(r"(PAX\d+)(\d{2,3})", check_label)
        if m:
            prefix = m.group(1)
            pin_num = int(m.group(2))
            max_pin = pax_max_pins.get(prefix, pin_num) if pax_max_pins else pin_num
            if max_pin > 5 and pin_num > math.ceil(max_pin / 2.0):
                is_first_half = False

        if is_first_half:
            label_cx = cx + LABEL_OFFSET_X_PAX_FIRST_HALF
            # Lower pins sit below the ring: offset 0 => top edge on the bottom circumference (cy + radius)
            anchor_y = (cy + radius) + LABEL_OFFSET_Y_PAX_FIRST_HALF
            align_v = "top"
        else:
            label_cx = cx + LABEL_OFFSET_X_PAX_SECOND_HALF
            # Upper pins sit above the ring: offset 0 => bottom edge on the top circumference (cy - radius)
            anchor_y = (cy - radius) - LABEL_OFFSET_Y_PAX_SECOND_HALF
            align_v = "bottom"

        label_cx += manual_dx
        anchor_y += manual_dy
        best_rect = make_label_rect(label_cx - (text_w/2), label_cx + (text_w/2), anchor_y, text_h, align_v)
        occupied_labels.append(best_rect)

    else:
        # Non-PAX (Test Point): attach the label to a chosen side of the ring; offset 0 => touches ring.
        attach = (LABEL_ATTACH_TP or "top").strip().lower()
        if attach == "bottom":
            # below ring, top edge on the bottom circumference (cy + radius); offset 0 => touches, positive => away (down)
            label_cx = cx + LABEL_OFFSET_X_TP + manual_dx
            anchor_y = (cy + radius) + LABEL_OFFSET_Y_TP + manual_dy
            best_rect = make_label_rect(label_cx - (text_w/2), label_cx + (text_w/2), anchor_y, text_h, "top")
        elif attach == "left":
            # left of ring, horizontally aligned right (right edge on the left circumference)
            center_y = cy + LABEL_OFFSET_Y_TP + manual_dy
            x1 = (cx - radius) - LABEL_OFFSET_X_TP + manual_dx
            x0 = x1 - text_w
            best_rect = fitz.Rect(x0, center_y - (text_h / 2.0), x1, center_y + (text_h / 2.0))
            label_cx = (x0 + x1) / 2.0
        elif attach == "right":
            # right of ring, horizontally aligned left (left edge on the right circumference)
            center_y = cy + LABEL_OFFSET_Y_TP + manual_dy
            x0 = (cx + radius) + LABEL_OFFSET_X_TP + manual_dx
            x1 = x0 + text_w
            best_rect = fitz.Rect(x0, center_y - (text_h / 2.0), x1, center_y + (text_h / 2.0))
            label_cx = (x0 + x1) / 2.0
        else:  # "top": above ring, vertically aligned bottom (bottom edge on the top circumference)
            label_cx = cx + LABEL_OFFSET_X_TP + manual_dx
            anchor_y = (cy - radius) - LABEL_OFFSET_Y_TP + manual_dy
            best_rect = make_label_rect(label_cx - (text_w/2), label_cx + (text_w/2), anchor_y, text_h, "bottom")

        occupied_labels.append(best_rect)
        draw_leader_line(page, label_cx, (best_rect.y0 + best_rect.y1) / 2.0, cx, cy, radius, RED)


    y0 = best_rect.y0
    if LABEL_BG_ENABLED:
        bg_color = get_label_bg_color()
        bg_rect = fitz.Rect(
            best_rect.x0 - LABEL_BG_PADDING_X,
            best_rect.y0 - LABEL_BG_PADDING_Y,
            best_rect.x1 + LABEL_BG_PADDING_X,
            best_rect.y1 + LABEL_BG_PADDING_Y,
        )
        page.draw_rect(bg_rect, color=bg_color, fill=bg_color, width=0, fill_opacity=LABEL_BG_ALPHA, overlay=True)

    for idx, line in enumerate(lines):
        baseline_y = y0 + (idx + 1) * line_step - (line_h * 0.2)
        line_w = fitz.get_text_length(line, fontname=LABEL_FONT_NAME, fontsize=LABEL_FONT_SIZE)
        line_x = label_cx - (line_w / 2.0)
        
        page.insert_text(
            fitz.Point(line_x, baseline_y), line,
            fontsize=LABEL_FONT_SIZE, fontname=LABEL_FONT_NAME, color=RED
        )
        
    return best_rect


# ============================================================
# PAGE AND FILE EXPORTING
# ============================================================

def circle_page(page: fitz.Page, pairs: list[tuple[str, str]], manual_offsets: dict = None) -> dict:
    if manual_offsets is None:
        manual_offsets = {}
        
    stats = {
        "total": len(pairs),
        "circled": 0,
        "not_found": 0,
        "locations": {},
        "not_found_labels": [],
        "resolved_aliases": {},
        "status": "SUCCESS",
        "debug_logs": []
    }
    occupied_labels: list[fitz.Rect] = []
    page_words = page.get_text("words")
    page_text = page.get_text("text").upper()

    # Precompute once per page (shared across all labels) instead of per-label
    cleaned_words = [(_clean(w[4]), w) for w in page_words]
    merged_words = _merge_words(page_words)

    dominant_prefix, prefix_counts = detect_dominant_prefix(page_words, page_text)
    stats["debug_logs"].append(f"PDF Format Scan: {prefix_counts} => Selected Prefix: '{dominant_prefix}'")

    pax_max_pins = {}
    for pdf_label, _ in pairs:
        norm = _normalize_label_for_visuals(pdf_label)
        m = re.fullmatch(r"(PAX\d+)(\d{2,3})", norm)
        if m:
            prefix = m.group(1)
            pin = int(m.group(2))
            if prefix not in pax_max_pins or pin > pax_max_pins[prefix]:
                pax_max_pins[prefix] = pin

    for pdf_label, net_label in pairs:
        aliases = build_aliases(pdf_label, dominant_prefix)
        stats["resolved_aliases"][pdf_label] = aliases

        rect = find_tp_rect(page, pdf_label, page_words, dominant_prefix,
                            aliases=aliases, cleaned_words=cleaned_words, merged_words=merged_words)

        if rect is None:
            stats["not_found"] += 1
            stats["not_found_labels"].append(pdf_label)
            stats["debug_logs"].append(f"❌ '{pdf_label}' NOT FOUND in PDF.")
            continue
        else:
            stats["debug_logs"].append(f"✅ '{pdf_label}' FOUND at Rect({round(rect.x0, 1)}, {round(rect.y0, 1)}).")

        cx, cy, radius, _ = circle_test_point(page, rect, pdf_label)
        
        mo = manual_offsets.get(pdf_label, {})
        mdx = float(mo.get("dx", 0.0))
        mdy = float(mo.get("dy", 0.0))
        
        label_rect = draw_net_label_top_touch(
            page, net_label, pdf_label, cx, cy, radius, 
            occupied_labels, pax_max_pins=pax_max_pins,
            manual_dx=mdx, manual_dy=mdy
        )

        stats["locations"][pdf_label] = {
            "cx": cx,
            "cy": cy,
            "radius": radius,
            "label_rect": [label_rect.x0, label_rect.y0, label_rect.x1, label_rect.y1],
            "aliases": aliases
        }
        stats["circled"] += 1

    if stats["total"] == 0:
        stats["status"] = "NO_DATA"
    elif stats["not_found"] == 0:
        stats["status"] = "SUCCESS"
    elif stats["circled"] > 0:
        stats["status"] = "PARTIAL_SUCCESS"
    else:
        stats["status"] = "FAILURE"

    return stats


def clear_all_annotations(doc: fitz.Document) -> int:
    removed = 0
    for page in doc:
        annot = page.first_annot
        while annot is not None:
            nxt = annot.next
            page.delete_annot(annot)
            removed += 1
            annot = nxt
    return removed


def annotate_pdf_file(
    input_pdf: str, output_pdf: str, pairs: list[tuple[str, str]],
    page_idx: int, clear_existing: bool = True,
    manual_offsets: dict = None
) -> dict:
    doc = fitz.open(input_pdf)

    if page_idx < 0 or page_idx >= len(doc):
        doc.close()
        raise ValueError(f"Invalid page index {page_idx}. PDF has {len(doc)} pages.")

    if clear_existing:
        clear_all_annotations(doc)

    page = doc[page_idx]
    stats = circle_page(page, pairs, manual_offsets=manual_offsets)

    out_dir = os.path.dirname(output_pdf) or "."
    os.makedirs(out_dir, exist_ok=True)

    fd, temp_out = tempfile.mkstemp(suffix=".pdf")
    os.close(fd)

    try:
        doc.save(temp_out, garbage=4, deflate=True)
        doc.close()
        if not os.path.exists(temp_out) or os.path.getsize(temp_out) <= 0:
            raise RuntimeError("PDF save failed")
        if os.path.exists(output_pdf):
            os.remove(output_pdf)
        shutil.move(temp_out, output_pdf)
        if not os.path.exists(output_pdf) or os.path.getsize(output_pdf) <= 0:
            raise RuntimeError("Final PDF output failed")
    finally:
        try:
            if not doc.is_closed:
                doc.close()
        except Exception:
            pass
        if os.path.exists(temp_out):
            try:
                os.remove(temp_out)
            except Exception:
                pass

    return stats